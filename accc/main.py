# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


import os
import sys
import traceback

name = "ll_hw"
note = "fnls"
note2 = "1000-gen"

pssepy_PATH=(r"""C:\Program Files\PTI\PSSE35\35.3\PSSPY39""")
sys.path.append(pssepy_PATH)

import psse35
import psspy
import pssexcel
import logging
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import csv
import pandas as pd
from psspy import _i, _f
import numpy as np

logger = logging.getLogger(__name__)

psspy.psseinit()
psspy.case("C:/nordics-2045/psse/" + name + "/nordics.sav")


def run_accc(name, note):

    #creates dfax file
    #psspy.dfax_2([0,1,0],r"C:/nordics-2045/psse/" + note + ".sub", r"C:/nordics-2045/psse/nordics.mon", r"C:/nordics-2045/psse/nordics.con", r"C:/nordics-2045/psse/" + note + ".dfx")

    #runs accc
    psspy.accc_with_dsp_3( 0.1,[0,0,0,1,1,1,0,0,0,0,0],"", r'C:/nordics-2045/psse/' + name + '/all_bz.dfx', r"C:/nordics-2045/psse/" + name + "/nordics" + name + ".acc","","","")

    accfile = r"C:/nordics-2045/psse/" + name + "/nordics" + name + ".acc"

    pssexcel.accc(accfile, ['s', 'b'], colabel='',
    stype='contingency', busmsm=0.5, sysmsm=5.0, rating='a', namesplit=True,
    xlsfile=name + note + ".xlsx", sheet='', overwritesheet=True, show=True, ratecon='b',
     baseflowvio=True,
    basevoltvio=True, flowlimit=100.0, flowchange=0.0, voltchange=0.0,
     swdrating='a',
    swdratecon='b',baseswdflowvio=False,basenodevoltvio=False,
     overloadreport=False)


def sort_accc_data(name, note):

    # read Excel sheet into a pandas data frame
    df = pd.read_excel(r"C:/nordics-2045/accc/" + name + note + ".xlsx")
    # concatenate columns A to F into a new column called "New_Column" and insert it after column F
    new_column = df.iloc[:, 0:7].apply(lambda x: ''.join(str(val) for val in x), axis=1)
    df.insert(7, 'Merged branch name', new_column)
    df.columns = ['Bus0', 'From_name', 'Voltage_0', 'Bus1', 'To_name', 'Voltage_1', 'Branch_id', 'Merged branch name', 'Contingency', 'MVAFLOW', 'AMPFLOW','RATE', '% FLOW', 'INFO']

    df = df.dropna(subset=['Merged branch name', '% FLOW'])

    # Add a new column to the DataFrame by multiplying '% FLOW' and 'rate rate1/rate1' columns
    df['Available MW'] = (100 - df['% FLOW']) * df['RATE'] * 0.01

    df_sorted = df.sort_values(by=['Merged branch name'])
    df_max = df.loc[df.groupby('Merged branch name')['% FLOW'].idxmax()]

    with pd.ExcelWriter(r"C:/nordics-2045/psse/" + name + "/ACCC_results" + name + note + ".xlsx") as writer:
        df_sorted.to_excel(writer, sheet_name='Branch FLow merged', index=False)
        df_max.to_excel(writer, sheet_name='Limiting branch', index=False)


def pick_out_branches(name, note):
    # Read the Excel sheet into a pandas data frame
    excel_file = pd.ExcelFile(r"C:/nordics-2045/psse/" + name + "/ACCC_results" + name + note + ".xlsx")

    # Load the existing sheets into data frames
    df_sorted = excel_file.parse(sheet_name='Branch FLow merged')

    # Apply the filtering conditions to the 'Limiting branch' sheet
    df_max = excel_file.parse(sheet_name='Limiting branch')
    df_filtered = df_max[df_max['Available MW'] < 600]
    df_filtered_400 = df_filtered[df_filtered['Voltage_0'] == 400]

    # Write the modified data frames to the output file along with existing sheets
    with pd.ExcelWriter(r"C:/nordics-2045/psse/" + name + "/ACCC_results" + name + note + ".xlsx") as writer:
        df_sorted.to_excel(writer, sheet_name='Branch FLow merged', index=False)
        df_max.to_excel(writer, sheet_name='Limiting branch', index=False)
        df_filtered.to_excel(writer, sheet_name='Heavy loaded branches', index=False)
        df_filtered_400.to_excel(writer, sheet_name='400 kV branches', index=False)

def run_sensitivity(name, note, note2):

    sensitivity_branches = pd.read_excel(r"C:/nordics-2045/psse/" + name + "/ACCC_results" + name + note + ".xlsx", sheet_name='400 kV branches')
    wb = openpyxl.Workbook()

    rad = 0

    for index, row in sensitivity_branches.iterrows():
        bus0 = row['Bus0']
        bus1 = row['Bus1']
        branch_id = str(row['Branch_id'])
        rad_str = str(index)
        rad = index

        psspy.report_output(2, f"C:/nordics-2045/psse/{name}/results/results" + rad_str + note2 +".txt", [0, 0])

        psspy.sensitivity_flow([bus0, bus1, 0, 1, 2], [1, 0, 0, 0, 0, 0, 0, 2, 1], [0.5, 0.1], branch_id,
                               [r"""ALL_BZ""", r"""RESERV"""], r'C:/nordics-2045/psse/' + name + '/all_bz_with_reserve.dfx')

    wb.save(f"C:/nordics-2045/psse/{name}/results.xlsx")
    sens_to_excel(name, rad, note, note2)


def sens_to_excel(name, rad, note, note2):
    # define the regular expression pattern to match the headline
    headline_pattern = r"SENSITIVITY FACTORS OF BRANCH FLOW \(MW\) ON\s+(.*?)\s+"

    # initialize the workbook and worksheet objects
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    df_branches = pd.read_excel(r"C:/nordics-2045/psse/" + name + "/ACCC_results" + name + note + ".xlsx", sheet_name='400 kV branches')

    # write the header row to the worksheet
    worksheet.append(["FILE NAME", "BUS NAME", "MW-SENS", "MVAFLOW", "RATE", "AVAILABLE MW"])

    # loop over the files
    for i in range(rad+1):
        # define the input and output file paths
        input_file = f"C:/nordics-2045/psse/{name}/results/results" + str(i) + note2 + ".txt"

        # initialize the list of bus names and mw-sens factors
        bus_names = []
        mw_sens_factors = []
        mva_flow = df_branches.loc[i, 'MVAFLOW']
        rate = df_branches.loc[i,'RATE']
        available_mw = df_branches.loc[i, 'Available MW']

        # read the input file and extract the bus names and mw-sens factors
        with open(input_file, "r") as f:
            # read the input file line by line
            for line in f:
                # check if the line matches the headline pattern
                match = re.match(headline_pattern, line)
                if match:
                    # if the line matches the headline pattern, extract the headline
                    headline = line.strip()
                    worksheet.append([headline]) # Write headline to worksheet
                else:
                    # if the line does not match the headline pattern, extract the bus name and mw-sens factor
                    tokens = line.strip().split()
                    if len(tokens) >= 2:
                        bus_name = " ".join(tokens[:-1])
                        mw_sens_factor = tokens[-1]
                        # append the file name, bus name, and mw-sens factor to the corresponding lists
                        bus_names.append(bus_name)
                        mw_sens_factors.append(mw_sens_factor)
                        # write the data to the worksheet
                        worksheet.append([f"results{i}.txt", bus_name, mw_sens_factor, mva_flow, rate, available_mw])

        # clear the lists of bus names and mw-sens factors for the next file
        bus_names = []
        mw_sens_factors = []

    # save the workbook to a file
    workbook.save(f"C:/nordics-2045/psse/{name}/results{name}{note}{note2}.xlsx")


def identify_buses(name, note, note2):

    df = pd.read_excel(f"C:/nordics-2045/psse/{name}/results{name}{note}{note2}.xlsx")

    df['MW-SENS'] = pd.to_numeric(df['MW-SENS'], errors='coerce')
    df['MVAFLOW'] = pd.to_numeric(df['MVAFLOW'], errors='coerce')
    df['RATE'] = pd.to_numeric(df['RATE'], errors='coerce')
    # df['AVAILABLE MW'] = pd.to_numeric(df['AVAILABLE MW'], errors='coerce')


    df['RESULTS'] = df['RATE'] - (np.where(df['MVAFLOW'] > 0, df['MW-SENS'] * -1, df['MW-SENS']) * -1000 + df['MVAFLOW']).abs()
    #df['RESULTS'] = df['RATE'] - (df['MW-SENS'] * 1000 + df['MVAFLOW']).abs()
    #df['AIDING(+)/INTERFERING(-)'] = df['RESULTS'] - df['AVAILABLE MW']
    df['AIDING(+)/INTERFERING(-)'] = df['RESULTS']
    print(df)
    # Write the results column to the Excel file
    with pd.ExcelWriter(f"C:/nordics-2045/psse/{name}/results{name}{note}{note2}.xlsx", mode='a',
                        engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sensistivity results', index=False)

    # Find the bus names with negative results
    # buses_less_than_zero = set(df.loc[df['AIDING(+)/INTERFERING(-)'] < 0, 'BUS NAME'].unique())
    # buses_more_than_zero = set(df.loc[df['AIDING(+)/INTERFERING(-)'] > 0, 'BUS NAME'].unique())

    buses_more_than_zero = set()
    buses_less_than_zero = set()

    # Iterate over the rows of the dataset
    for index, row in df.iterrows():
        aid = row['AIDING(+)/INTERFERING(-)']
        bus = row['BUS NAME']

        if aid > 0:
            buses_more_than_zero.add(bus)
        elif aid < 0:
            buses_less_than_zero.add(bus)

    # Exclude buses present in buses_less_than_zero from buses_more_than_zero
    buses_only_good = buses_more_than_zero.difference(buses_less_than_zero)
    buses_only_bad = buses_less_than_zero.difference(buses_more_than_zero)

    buses_both = buses_more_than_zero.intersection(buses_less_than_zero)

    # Load the Excel file again and create a new sheet for the results
    workbook = openpyxl.load_workbook(f"C:/nordics-2045/psse/{name}/results{name}{note}{note2}.xlsx")
    worksheet = workbook.create_sheet(title='Buses violating-non violating')

    # Write the header row to the new sheet
    worksheet.append(["Bus - violating", 'Bus - non violating', 'Bus present in both'])

    # Loop through the bus names and write them to the new sheet
    for i, bus_name in enumerate(buses_only_bad):
        worksheet.cell(row=i + 2, column=1, value=bus_name)

    for i, bus_name in enumerate(buses_only_good):
        worksheet.cell(row=i + 2, column=2, value=bus_name)

    for i, bus_name in enumerate(buses_both):
        worksheet.cell(row=i + 2, column=3, value=bus_name)

    # Save the changes to the Excel file
    workbook.save(f"C:/nordics-2045/psse/{name}/results{name}{note}{note2}.xlsx")


run_accc(name, note)
sort_accc_data(name, note)
#pick_out_branches(name, note)
#run_sensitivity(name, note, note2)
#identify_buses(name, note, note2)

